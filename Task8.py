import openpyxl
from datetime import datetime, timedelta

def main():
    file_path = 'C:/Users/Kamal Teja INT-212/hello-docker/att.xlsx'  # Update with your Excel file path
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    today = datetime.now().strftime('%b %d')
    prev_6_days = [(datetime.now() - timedelta(days=i)).strftime('%b %d') for i in range(1, 7)]

    wfh_today = []
    wfo_today = []
    wfh_prev_days = []
    wfo_prev_days = []
    all_employees = set()
    filled_employees = set()
    missing_entries = []

    missing_entries_by_date = {}  # Initialize the dictionary

    header_row = next(sheet.iter_rows(min_row=1, values_only=True))
    date_columns = header_row[1:]

    for row in sheet.iter_rows(min_row=2, values_only=True):
        emp_id = row[0]
        all_employees.add(emp_id)

        for day_index, status in enumerate(row[1:], start=1):
            date = date_columns[day_index - 1]
            
            if date == today:
                if status == 'WFH':
                    wfh_today.append(emp_id)
                elif status == 'WFO':
                    wfo_today.append(emp_id)
            
            if date in prev_6_days:
                if status == 'WFH':
                    wfh_prev_days.append(emp_id)
                elif status == 'WFO':
                    wfo_prev_days.append(emp_id)

            if status in ['WFH', 'WFO']:
                filled_employees.add(emp_id)
        
            if status is None:
                if date not in missing_entries_by_date:
                    missing_entries_by_date[date] =set()
                missing_entries_by_date[date].add(emp_id)

        filled_dates = [date for day_index, date in enumerate(date_columns) if row[day_index + 1] in ['WFH', 'WFO']]
        if len(filled_dates) != 6:
            missing_entries.append(emp_id)

    print("WFH Count for Today:", len(wfh_today))
    print("WFO Count for Today:", len(wfo_today))
    print("WFH Count for Previous 6 Days:", len(wfh_prev_days))
    print("WFO Count for Previous 6 Days:", len(wfo_prev_days))
    print("\nEmployees with Missing Attendance for Each Date:")
    for date, missing_employees in missing_entries_by_date.items():
        print("-" * 30)
        print(f"For {date}: {', '.join(map(str, missing_employees))}")

if __name__ == "__main__":
    main()
